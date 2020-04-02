from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(filename= 'A_sample.xlsx')
print(workbook.sheetnames)
sheet = workbook.active
print(sheet['A1'].value)